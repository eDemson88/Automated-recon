# ----------------------------------------------------------------------
# 1️⃣ IMPORT & UTILS
# ----------------------------------------------------------------------
import argparse

import openpyxl
from pathlib import Path
from typing import List, Dict, Any
from openpyxl.utils import column_index_from_string

# ----------------------------------------------------------------------
# 2️⃣ LOAD & KONVERSI SHEET
# ----------------------------------------------------------------------
def load_sheet(path: Path, sheet_name: str = None) -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Buka workbook dan kembalikan worksheet.
    Jika sheet_name = None → gunakan sheet aktif.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return wb, ws

def sheet_to_records(ws) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Konversi seluruh sheet menjadi:
        header  – list nama kolom (baris pertama)
        records – list of dict, satu dict per baris data
    Nilai None untuk sel kosong.
    """
    rows = list(ws.iter_rows(values_only=True))
    header = [str(cell).strip() for cell in rows[0]]
    records = [{header[i]: row[i] for i in range(len(header))} for row in rows[1:]]
    return header, records

def build_lookup(records: List[Dict[str, Any]], key: str) -> Dict[Any, List[int]]:
    """
    Membuat kamus { nilai → list indeks baris } untuk kolom tertentu.
    Nilai None atau 0 di‑abaikan.
    """
    lookup: Dict[Any, List[int]] = {}
    for idx, rec in enumerate(records):
        val = rec.get(key)
        if val is None or (isinstance(val, (int, float)) and val == 0):
            continue
        lookup.setdefault(val, []).append(idx)
    return lookup

# ----------------------------------------------------------------------
# 4️⃣ FUNGSI MENAMBAHKAN KOLUMN BALANCE
# ----------------------------------------------------------------------
def add_balance_column(
    ws,
    balance_col_letter: str = "H",   # kolom Balance tempat hasil ditulis
    debit_col_letter: str = "F",     # kolom Debit (ubah bila berbeda)
    credit_col_letter: str = "G",    # kolom Credit (ubah bila berbeda)
    opening_balance_row: int = 2,    # baris yang menyimpan Opening Balance (tidak diubah)
    first_data_row: int = 3,         # baris pertama yang **dihitung** (biasanya 3)
) -> None:
    """
    Mengisi kolom Balance dengan rumus:
        Balance[i] = Balance[i‑1] – Debit[i] + Credit[i]

    - Nilai pada `opening_balance_row` (mis. C2) dipertahankan dan dipakai
      sebagai nilai awal.
    - Perhitungan dimulai pada `first_data_row` (default = 3) – tepat
      setelah baris Opening Balance.
    - Jika sel Debit atau Credit kosong atau bukan angka, dianggap 0.
    """
    # Konversi huruf kolom → indeks 1‑based
    bal_idx   = column_index_from_string(balance_col_letter)
    debit_idx = column_index_from_string(debit_col_letter)
    credit_idx = column_index_from_string(credit_col_letter)

    # -----------------------------------------------------------------
    # 1️⃣ Baca Opening Balance (baris 2) – gunakan 0 bila kosong
    # -----------------------------------------------------------------
    opening_cell = ws.cell(row=opening_balance_row, column=bal_idx)
    try:
        previous_balance = (
            float(opening_cell.value) if opening_cell.value is not None else 0.0
        )
    except (ValueError, TypeError):
        previous_balance = 0.0

    # Pastikan nilainya tetap numerik
    opening_cell.value = previous_balance

    # -----------------------------------------------------------------
    # 2️⃣ Loop mulai dari baris pertama data yang akan dihitung
    # -----------------------------------------------------------------
    for row in range(first_data_row, ws.max_row + 1):
        debit_cell  = ws.cell(row=row, column=debit_idx)
        credit_cell = ws.cell(row=row, column=credit_idx)

        # Debit → 0 bila kosong / tak dapat dikonversi
        try:
            debit_val = (
                float(debit_cell.value) if debit_cell.value is not None else 0.0
            )
        except (ValueError, TypeError):
            debit_val = 0.0

        # Credit → 0 bila kosong / tak dapat dikonversi
        try:
            credit_val = (
                float(credit_cell.value) if credit_cell.value is not None else 0.0
            )
        except (ValueError, TypeError):
            credit_val = 0.0

        # Hitung balance
        current_balance = previous_balance - debit_val + credit_val

        # Tulis ke kolom Balance
        ws.cell(row=row, column=bal_idx, value=current_balance)

        # Simpan untuk iterasi selanjutnya
        previous_balance = current_balance

# ----------------------------------------------------------------------
# 3️⃣ REKONSILIASI – PAIRING & PEMISAHAN UNMATCHED
# ----------------------------------------------------------------------
def reconcile_with_separate_unmatched(
    src_path: Path,
    output_path: Path,
    sheet_name: str = None,
    debit_col: str = "Debit",
    credit_col: str = "Credit",
    dry_run: bool = False,
) -> None:
    # ---------- 1. Load ----------
    wb_src, ws_src = load_sheet(src_path, sheet_name)
    header, records = sheet_to_records(ws_src)

    # ---------- 2. Pairing ----------
    debit_lookup  = build_lookup(records, debit_col)
    credit_lookup = build_lookup(records, credit_col)

    used = set()
    paired_rows: List[int] = []          # urutan indeks (0‑based) yang sudah dipasangkan
    unmatched_debits: List[int] = []     # debit‑only
    unmatched_credits: List[int] = []    # credit‑only

    for i, rec in enumerate(records):
        if i in used:
            continue

        d_val = rec.get(debit_col)
        c_val = rec.get(credit_col)

        # Debit → cari credit yang sama
        if d_val not in (None, 0) and d_val in credit_lookup:
            cand = [idx for idx in credit_lookup[d_val] if idx not in used]
            if cand:
                j = cand[0]
                paired_rows.extend([j, i])      # credit dulu, debit di bawahnya
                used.update({j, i})
                continue

        # Credit → cari debit yang sama (jika credit muncul dulu)
        if c_val not in (None, 0) and c_val in debit_lookup:
            cand = [idx for idx in debit_lookup[c_val] if idx not in used]
            if cand:
                j = cand[0]
                paired_rows.extend([i, j])      # credit dulu, debit di bawahnya
                used.update({i, j})
                continue

        # Tidak ada pasangan
        if d_val not in (None, 0):
            unmatched_debits.append(i)
        else:
            unmatched_credits.append(i)
        used.add(i)

    # ---------- 3. Urutan akhir ----------
    final_order = ["OPENING"] + paired_rows + ["BLANK"] + unmatched_debits + unmatched_credits

    if dry_run:
        print("[DRY‑RUN] Contoh urutan indeks (termasuk 'BLANK'):")
        print(final_order[:20])
        return

    # ---------- 4. Tulis ke workbook baru ----------
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_src.title

    # 4a. Header (baris 1)
    for col_idx, col_name in enumerate(header, start=1):
        ws_out.cell(row=1, column=col_idx, value=col_name)

    # 4b. Data
    # Data (dengan penanganan marker)
    out_row = 2                                    # mulai menulis di baris 2
    for item in final_order:
        if item == "OPENING":
            # Salin nilai Opening Balance (H2) saja
            src_cell = ws_src.cell(row=2, column=column_index_from_string("H"))
            dst_cell = ws_out.cell(row=out_row, column=column_index_from_string("H"))
            dst_cell.value = src_cell.value
            # (optional) copy style here if you want
            out_row += 1
            continue

        if item == "BLANK":
            out_row += 1                           # satu baris kosong
            continue

        src_idx = item                              # indeks dalam `records`
        src_row = src_idx + 2                       # baris sumber (data dimulai pada baris 2)

        for col_idx, col_name in enumerate(header, start=1):
            value = records[src_idx].get(col_name)
            ws_out.cell(row=out_row, column=col_idx, value=value)

        out_row += 1

    # Simpan workbook hasil
    wb_out.save(output_path)
    print(f"✅ selesai – file output disimpan di: {output_path}")



# ----------------------------------------------------------------------
# 5️⃣ CLI / MAIN
# ----------------------------------------------------------------------
def build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Reconcile debit‑credit, pisahkan unmatched, dan tambahkan "
            "kolom Balance (mulai perhitungan dari baris 3)."
        )
    )

    parser.add_argument("-i", "--input", dest="input_file", type=Path, required=True,
                        help="Path ke file Excel sumber.")
    parser.add_argument("-o", "--output", dest="output_file", type=Path, required=True,
                        help="Path file hasil.")
    parser.add_argument("-s", "--sheet", dest="sheet_name", default=None,
                        help="Nama sheet (default: sheet aktif).")
    parser.add_argument("--debit-col", dest="debit_col", default="Debit",
                        help="Nama kolom Debit.")
    parser.add_argument("--credit-col", dest="credit_col", default="Credit",
                        help="Nama kolom Credit.")
    parser.add_argument("--balance-col", dest="balance_col", default="Balance",
                        help="Nama kolom Balance yang akan ditulis.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Hanya tampilkan urutan indeks, tidak menulis file.")
    return parser

def main():
    parser = build_cli()
    args = parser.parse_args()

    # -----------------------------------------------------------------
    # 1️⃣ Rekonsiliasi + pisahkan unmatched
    # -----------------------------------------------------------------
    reconcile_with_separate_unmatched(
        src_path=args.input_file,
        output_path=args.output_file,
        sheet_name=args.sheet_name,
        debit_col=args.debit_col,
        credit_col=args.credit_col,
        dry_run=args.dry_run,
    )

    if args.dry_run:
        return  # tidak lanjut ke step balance bila dry‑run

    # -----------------------------------------------------------------
    # 2️⃣ Buka kembali file output untuk menambahkan kolom Balance
    # -----------------------------------------------------------------
    wb_out, ws_out = load_sheet(args.output_file, args.sheet_name)

    # Jika kolom Balance belum ada, tambahkan satu kolom pada posisi yang diinginkan.
    # Di sini contoh menambahkan pada kolom C (Anda dapat ubah hurufnya bila perlu).
    add_balance_column(
        ws_out,
        balance_col_letter="H",   # Opening Balance sudah di H2
        debit_col_letter="F",     # sesuaikan dengan file Anda
        credit_col_letter="G",    # sesuaikan dengan file Anda
        opening_balance_row=2,    # H2 (tidak diubah)
        first_data_row=3          # perhitungan dimulai di baris 3
    )

    # Simpan kembali workbook yang sudah berisi kolom Balance
    wb_out.save(args.output_file)
    print(f"✅ Semua selesai – file akhir dengan Balance ada di: {args.output_file}")

if __name__ == "__main__":
    main()