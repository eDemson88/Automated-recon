# --------------------------------------------------------------
# reconcile.py
# --------------------------------------------------------------
import sys
import statistics
from pathlib import Path

import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, PatternFill, Fill

# ----------------------------------------------------------------------
# 1️⃣  Load workbook (data_only=True keeps only the displayed values)
# ----------------------------------------------------------------------
def load_workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True)

# ----------------------------------------------------------------------
# 2️⃣  Header‑style helper – fixes StyleProxy problem
# ----------------------------------------------------------------------
def get_header_style(ws):
    """
    Copies the visual style of the first header cell (A1) into a NamedStyle.
    Works with the openpyxl version you have (no `theme` on Font, no
    `justification` on Alignment, and Fill must be a concrete Fill object).
    The style is registered with the workbook so you can apply it by name.
    """
    first = ws["A1"]

    # ---- Font ---------------------------------------------------------
    header_font = Font(
        name=first.font.name,
        sz=first.font.sz,
        bold=first.font.bold,
        italic=first.font.italic,
        vertAlign=first.font.vertAlign,
        underline=first.font.underline,
        strike=first.font.strike,
        color=first.font.color,
        scheme=first.font.scheme,
        family=first.font.family,
        charset=first.font.charset,
    )

    # ---- Alignment ----------------------------------------------------
    # Only primitive attributes are copied – no StyleProxy objects.
    align_kwargs = {
        "horizontal":   first.alignment.horizontal,
        "vertical":     first.alignment.vertical,
        "textRotation": first.alignment.textRotation,
        "wrapText":     first.alignment.wrapText,
        "shrinkToFit":  first.alignment.shrinkToFit,
        "indent":       first.alignment.indent,
        "readingOrder": first.alignment.readingOrder,
    }
    align_kwargs = {k: v for k, v in align_kwargs.items() if v is not None}
    header_alignment = Alignment(**align_kwargs)

    # ---- Border -------------------------------------------------------
    header_border = Border(
        left=first.border.left,
        right=first.border.right,
        top=first.border.top,
        bottom=first.border.bottom,
        diagonal=first.border.diagonal,
        diagonal_direction=first.border.diagonal_direction,
        outline=first.border.outline,
        vertical=first.border.vertical,
        horizontal=first.border.horizontal,
    )

    # ---- Fill ---------------------------------------------------------
    src = first.fill
    if isinstance(src, PatternFill):
        header_fill = PatternFill(
            fill_type=src.fill_type,
            start_color=src.start_color,
            end_color=src.end_color,
            fgColor=src.fgColor,
            bgColor=src.bgColor,
        )
    else:
        header_fill = Fill()                     # fallback – empty fill

    # ---- Assemble NamedStyle -------------------------------------------
    header_style = NamedStyle(name="header_style")
    header_style.font = header_font
    header_style.alignment = header_alignment
    header_style.border = header_border
    header_style.fill = header_fill

    # Register the style so it can be referenced later by name
    ws.parent.add_named_style(header_style)
    return header_style

# ----------------------------------------------------------------------
# 3️⃣  Simple risk heuristic (unchanged)
# ----------------------------------------------------------------------
def compute_risk(row, median_amount):
    amount = row["Debit"] - row["Credit"]
    large = abs(amount) > 3 * median_amount
    unmatched = row["Net"] != 0
    score = (0.6 if large else 0) + (0.5 if unmatched else 0)
    return min(score, 1.0)

# ----------------------------------------------------------------------
# 4️⃣  Move‑a‑row helper (preserves styles, comments, merges, etc.)
# ----------------------------------------------------------------------
def move_row(ws, src_idx: int, dst_idx: int) -> None:
    """Move an entire row from src_idx to dst_idx (1‑based)."""
    if src_idx == dst_idx:
        return

    max_col = ws.max_column

    # --- 1️⃣ copy source cells (keep objects, not just values) -----------
    src_cells = [ws.cell(row=src_idx, column=c) for c in range(1, max_col + 1)]

    # --- 2️⃣ delete the source row (rows below shift up) -----------------
    ws.delete_rows(src_idx)

    # If source was above destination, the destination index moved up by 1
    if src_idx < dst_idx:
        dst_idx -= 1

    # --- 3️⃣ insert a blank row at the destination -----------------------
    ws.insert_rows(dst_idx)

    # --- 4️⃣ paste the saved cells into the new location -----------------
    for col, src in enumerate(src_cells, start=1):
        dst = ws.cell(row=dst_idx, column=col)
        dst.value = src.value
        if src.comment:
            dst.comment = src.comment
        if src.hyperlink:
            dst.hyperlink = src.hyperlink

    # --- 5️⃣ fix merged cells that referenced the moved row -------------
    from openpyxl.utils import range_boundaries
    new_merged = []
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        if min_r <= src_idx <= max_r:               # the moved row was inside
            offset = dst_idx - src_idx
            min_r += offset
            max_r += offset
        new_merged.append(
            f"{ws.cell(row=min_r, column=min_c).coordinate}:"
            f"{ws.cell(row=max_r, column=max_c).coordinate}"
        )
    ws.merged_cells.ranges = []          # clear old list
    for rng in new_merged:
        ws.merge_cells(rng)

# ----------------------------------------------------------------------
# 5️⃣  Main routine – loads, processes, moves rows, adds columns, saves
# ----------------------------------------------------------------------
def main(input_path: str, output_path: str):
    wb = load_workbook(Path(input_path))
    ws = wb.active

    # Preserve the original header styling
    header_style = get_header_style(ws)

    # --------------------------------------------------------------
    # Identify columns (case‑insensitive)
    # --------------------------------------------------------------
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name.lower(): i for i, name in enumerate(header) if name}
    required = ["date", "account", "debit", "credit", "description"]
    for r in required:
        if r not in col_idx:
            sys.exit(f"Missing required column: {r}")

    # --------------------------------------------------------------
    # Read rows – keep the original worksheet row number for later moves
    # --------------------------------------------------------------
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=False):
        row = {
            "Date": r[col_idx["date"]].value,
            "Account": r[col_idx["account"]].value,
            "Debit": float(r[col_idx["debit"]].value or 0),
            "Credit": float(r[col_idx["credit"]].value or 0),
            "Description": r[col_idx["description"]].value,
            "RawCells": r,                 # reference to the original cell objects
            "RowIdx": r[0].row,            # current worksheet row number
        }
        rows.append(row)

    # --------------------------------------------------------------
    # Compute per‑account net totals (used for risk scoring)
    # --------------------------------------------------------------
    account_totals = {}
    for row in rows:
        net = row["Debit"] - row["Credit"]
        account_totals.setdefault(row["Account"], 0.0)
        account_totals[row["Account"]] += net

    # --------------------------------------------------------------
    # Add reconciliation flag, net per row, risk score
    # --------------------------------------------------------------
    amounts = [abs(r["Debit"] - r["Credit"]) for r in rows if (r["Debit"] or r["Credit"])]
    median_amount = statistics.median(amounts) if amounts else 0

    for row in rows:
        row["Net"] = account_totals[row["Account"]]
        row["Reconciled"] = "YES" if abs(row["Net"]) < 1e-6 else "NO"
        row["RiskScore"] = compute_risk(row, median_amount)

    # ------------------------------------------------------------------
    # 6️⃣  MOVE EVERY DEBIT ROW DIRECTLY BELOW ITS MATCHING CREDIT ROW
    # ------------------------------------------------------------------
    #
    # Matching rule:
    #   * Same Account
    #   * Same absolute amount (Credit value == Debit value)
    #   * Credit > 0, Debit > 0
    #
    # The algorithm scans the sheet **top‑to‑bottom**, finds a credit row,
    # then looks *forward* for the first debit row that satisfies the rule.
    # When found, the debit row is moved to the row immediately after the credit.
    #
    processed_debits = set()          # avoid moving the same debit twice

    for i, credit in enumerate(rows):
        if credit["Credit"] <= 0:
            continue

        # look forward for a matching debit that hasn't been moved yet
        for j in range(i + 1, len(rows)):
            debit = rows[j]
            if (
                debit["Debit"] > 0
                and debit["Account"] == credit["Account"]
                and abs(debit["Debit"] - credit["Credit"]) < 1e-6
                and debit["RowIdx"] not in processed_debits
            ):
                src = debit["RowIdx"]
                dst = credit["RowIdx"] + 1

                if src != dst:
                    move_row(ws, src, dst)

                    # ---- Update stored RowIdx for all rows that shifted ----
                    # When a row moves up, rows between src and dst shift down 1,
                    # and vice‑versa.
                    shift = -1 if src > dst else 1
                    low, high = (dst, src) if src > dst else (src, dst)
                    for r in rows:
                        if low <= r["RowIdx"] < high:
                            r["RowIdx"] += shift
                    # finally set the moved debit to its new location
                    debit["RowIdx"] = dst

                processed_debits.add(debit["RowIdx"])
                break   # only move the *first* matching debit for this credit

    # ------------------------------------------------------------------
    # 7️⃣  Append new columns (Reconciled, RiskScore) with original header style
    # ------------------------------------------------------------------
    new_headers = ["Reconciled", "RiskScore"]
    start_col = len(header) + 1
    for offset, name in enumerate(new_headers):
        cell = ws.cell(row=1, column=start_col + offset, value=name)
        cell.style = header_style

    # Populate the new columns
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=start_col,     value=row["Reconciled"])
        ws.cell(row=r_idx, column=start_col + 1, value=row["RiskScore"])

    # ------------------------------------------------------------------
    # 8️⃣  Print risky rows (optional diagnostic)
    # ------------------------------------------------------------------
    for row in rows:
        if row["RiskScore"] > 0.7:
            print(
                f"Risk alert – Account {row['Account']} on {row['Date']}: "
                f"RiskScore={row['RiskScore']:.2f}. "
                f"Suggestion: review large unmatched transaction."
            )

    # ------------------------------------------------------------------
    # 9️⃣  Save the workbook
    # ------------------------------------------------------------------
    wb.save(output_path)
    print(f"Reconciliation completed. Output saved to {output_path}")

# ----------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("Usage: python reconcile.py INPUT.xlsx OUTPUT.xlsx")
    main(sys.argv[1], sys.argv[2])